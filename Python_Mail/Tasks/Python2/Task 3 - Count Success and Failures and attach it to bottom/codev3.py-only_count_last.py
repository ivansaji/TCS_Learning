#imports

# importing openpyxl module
import openpyxl as xl

#modules
def count_data():
    #method to count data# opening the source excel file
    filename ="Z:\\share\\Tasks\\Python2\\Task 3 - Count Success and Failures and attach it to bottom\\sheet.xlsx"
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
    ws2 = wb1.active

    # calculate total number of rows and 
    mr = ws1.max_row

    #calculate max number of columns in sheet
    mc=ws1.max_column

    suc = 0
    fail = 0
    total = 0
    
    for j in range (3, mr + 1):
        #row operation
        #reading each row value of a col
        c = ws1.cell(row = j, column = mc)
        if(str(c.value)=='S'):
            suc=suc+1
        elif(str(c.value)=='F'):
            fail=fail+1
    print("The Total is "+ str(suc+fail))
    print("The Success is "+ str(suc))
    print("The Fail is "+ str(fail))
    print("___________________________")

    #appending
    ws2.cell(row = mr-2, column = mc).value = suc
    ws2.cell(row = mr-1, column = mc).value = fail
    ws2.cell(row = mr, column = mc).value = int(suc+fail)
    
    #saving
    wb1.save(str(filename))


#main
count_data()