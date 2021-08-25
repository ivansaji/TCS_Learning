# importing openpyxl module
import openpyxl as xl
  
# opening the source excel file
filename ="Z:\\share\\Tasks\Python3\\Task 2 - Append Excel Operation\\sheeta.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]
  
# opening the destination excel file 
filename1 ="Z:\\share\\Tasks\Python3\\Task 2 - Append Excel Operation\\sheetb.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# calculate total number of rows and 
mr = ws1.max_row

#calculate max number of columns in sheet 2
mc2=ws2.max_column

#SrcCol is the required column number to copy
SrcCol=2

#DstCol is the destination column to paste
DstCol = mc2+1

# copying the cell values from source 
# excel file to destination excel file

for i in range(1,mr+1):
    # reading cell value from source excel file
    c = ws1.cell(row = i, column = SrcCol)

    # writing the read value to destination excel file
    ws2.cell(row = i, column = DstCol).value = c.value

# saving the destination excel file
wb2.save(str(filename1))