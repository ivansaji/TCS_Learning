#imports

# importing openpyxl module
import openpyxl as xl

#modules

def rip(c):
    x=c.split(".")
    return x[0]

def count_data():
    #method to count data# opening the source excel file
    filename ="Z:\\share\\Tasks\Python3\\Task 4 - Strip and append\\sheet.xlsx"
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
    ws2 = wb1.active

    # calculate total number of rows and 
    mr = ws1.max_row

    #calculate max number of columns in sheet
    mc=ws1.max_column

    SrcCol=1
    #mc+1 sets it to last col
    DstCol=mc+1

    for i in range(3,mr+1):
        #row operation

        #reading data
        c = ws1.cell(row = i, column = SrcCol)
        newdata=rip(str(c.value))

        #writing new data to destination col
        ws2.cell(row = i, column = DstCol).value = str(newdata)

    print("Task complete")
    #saving
    wb1.save(str(filename))


#main
count_data()