#imports

# importing openpyxl module
import openpyxl as xl

#modules


def del_row():
    #method to count data# opening the source excel file
    filename ="Z:\\share\\Tasks\\Python3\\Task 5 - Delete first N rows\\sheet.xlsx"
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
    ws2 = wb1.active

    #StrCol
    StrCol=1
    #mc+1 sets it to last col
    EndCol=4

    print("Maximum rows before removing:", ws2.max_row)

    ws2.delete_rows(StrCol,EndCol)

    print("Maximum rows after removing:", ws2.max_row)

    print("Task complete")
    #saving
    wb1.save(str(filename))


#main
del_row()