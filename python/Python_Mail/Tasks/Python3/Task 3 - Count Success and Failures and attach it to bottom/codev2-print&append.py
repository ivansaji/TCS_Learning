#imports

# importing openpyxl module
import openpyxl as xl

#modules
def count_data():
    #method to count data# opening the source excel file
    filename ="Z:\\share\\Tasks\\Python3\\Task 3 - Count Success and Failures and attach it to bottom\\sheet.xlsx"
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
    ws2 = wb1.active

    # calculate total number of rows and 
    mr = ws1.max_row

    #calculate max number of columns in sheet
    mc=ws1.max_column

    for i in range (2, mc + 1):
        suc = 0
        fail = 0
        total = 0
        #col Operation
        for j in range (3, mr + 1):
            #row operation
            #reading each row value of a col
            c = ws1.cell(row = j, column = i)
            if(str(c.value)=='S'):
                suc=suc+1
            elif(str(c.value)=='F'):
                fail=fail+1
        print("The Total is "+ str(total))
        print("The Success is "+ str(suc))
        print("The Fail is "+ str(fail))
        print("___________________________")

        #appending
        ws2.cell(row = mr+1, column = i).value = suc
        ws2.cell(row = mr+2, column = i).value = fail
        ws2.cell(row = mr+3, column = i).value = int(suc+fail)
    
    #Giving Headings
    ws2.cell(row = mr+1, column = 1).value = "Success"
    ws2.cell(row = mr+2, column = 1).value = "Failed"
    ws2.cell(row = mr+3, column = 1).value = "Total Count"
    #saving
    wb1.save(str(filename))


#main
count_data()